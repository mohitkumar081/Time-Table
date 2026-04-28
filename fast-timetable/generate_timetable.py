"""
FAST NUCES Timetable Generator - FIXED VERSION
Usage:  python generate_timetable.py
Output: data/timetable.json
"""
import openpyxl, json, os, sys, re
from datetime import datetime

XLSX_FILE   = 'FAST_NUCES_Timetable.xlsx'
OUTPUT_FILE = os.path.join('data', 'timetable.json')
VALID_DAYS  = ['MONDAY','TUESDAY','WEDNESDAY','THURSDAY','FRIDAY']
SKIP_ROWS   = {'classrooms','computing labs','engineering labs','venues/time','slots',''}
SKIP_STARTS = ['reserved','classrooms','computing','engineering']

TIME_SLOT_ORDER = {
    '08:00-8:50':1,'08:00-08:50':1,
    '08:55-09:45':2,
    '09:50:-10:40':3,'09:50-10:40':3,
    '10:45-11:35':4,
    '11:40-12:30':5,
    '12:35-1:25':6,
    '1:30-2:20':7,
    '2:25-3:15':8,
    '3:20-4:10':9,
}

def parse_course_section(line):
    m = re.search(r'\b(B[A-Z]{1,3}-\d+[A-Z])', line)
    if m:
        return line[:m.start()].strip(), line[m.start():].strip()
    parts = line.split(' ', 1)
    return parts[0].strip(), (parts[1].strip() if len(parts)>1 else '')

def main():
    xlsx_path = XLSX_FILE if os.path.exists(XLSX_FILE) else None
    if not xlsx_path:
        for f in os.listdir('.'):
            if f.endswith('.xlsx') and 'timetable' in f.lower():
                xlsx_path=f; break
    if not xlsx_path:
        print(f"ERROR: No xlsx found! Rename to: {XLSX_FILE}"); sys.exit(1)

    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)
    entries = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        day_raw = (ws.cell(1,1).value or '').strip().upper()
        if day_raw not in VALID_DAYS: continue
        day = day_raw

        slot_times = {}
        for c in range(1, ws.max_column+1):
            val = ws.cell(3,c).value
            if val and ':' in str(val):
                slot_times[c] = str(val).strip()

        print(f"  {day} ('{sheet_name}'): {len(slot_times)} slots")

        for r in range(4, ws.max_row+1):
            room_raw = ws.cell(r,1).value
            if not room_raw: continue
            room = str(room_raw).strip()
            if room.lower() in SKIP_ROWS: continue
            if any(room.lower().startswith(x) for x in SKIP_STARTS): continue

            for c, time in slot_times.items():
                cell_val = ws.cell(r,c).value
                if not cell_val: continue
                cell_str = str(cell_val).strip()
                if not cell_str or cell_str.lower()=='none': continue

                lines = [l.strip() for l in cell_str.replace('\r','').split('\n') if l.strip()]
                if not lines: continue

                courseCode, section = parse_course_section(lines[0])
                teacher = ' '.join(lines[1:]).strip().lstrip()
                slot_order = TIME_SLOT_ORDER.get(time, 99)

                entries.append({
                    'day': day, 'time': time, 'slot': slot_order,
                    'room': room, 'courseCode': courseCode,
                    'section': section, 'teacher': teacher,
                    'key': f"{day}|{time}|{room}|{courseCode}|{section}"
                })

    os.makedirs('data', exist_ok=True)
    with open(OUTPUT_FILE,'w',encoding='utf-8') as f:
        json.dump(entries,f,ensure_ascii=False,separators=(',',':'))

    from collections import Counter
    counts = Counter(e['day'] for e in entries)
    print(f"\nDone! {len(entries)} entries saved to {OUTPUT_FILE}")
    for d in VALID_DAYS:
        print(f"  {d}: {counts.get(d,0)}")

if __name__=='__main__':
    main()
